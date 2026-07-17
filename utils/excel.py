# ==========================================================
# SPORTAVALIA PRO V3
# utils/excel.py
# PARTE 1
# ==========================================================

from pathlib import Path
from openpyxl import load_workbook
import shutil

# ==========================================================
# PASTAS
# ==========================================================

BASE_DIR = Path(__file__).resolve().parent.parent

PASTA_MODELOS = BASE_DIR / "modelos"

PASTA_ATLETAS = BASE_DIR / "atletas"

MODELO = PASTA_MODELOS / "modelo_atleta.xlsx"

PASTA_ATLETAS.mkdir(exist_ok=True)

# ==========================================================
# ATLETAS
# ==========================================================

def listar_atletas():

    arquivos = sorted(

        PASTA_ATLETAS.glob("*.xlsx"),

        key=lambda arq: arq.stem.lower()

    )

    return [a.stem for a in arquivos]


def caminho_atleta(nome):

    return PASTA_ATLETAS / f"{nome}.xlsx"


def atleta_existe(nome):

    return caminho_atleta(nome).exists()


def copiar_modelo(nome):

    destino = caminho_atleta(nome)

    shutil.copy(MODELO, destino)

    return destino


def criar_atleta(nome):

    if atleta_existe(nome):

        return caminho_atleta(nome)

    return copiar_modelo(nome)


def abrir_atleta(nome):

    arquivo = caminho_atleta(nome)

    if not arquivo.exists():

        return None

    return load_workbook(arquivo)

# ==========================================================
# SALVAR CADASTRO
# ==========================================================

def salvar_cadastro(

    nome,
    sexo,
    idade,
    nascimento,
    esporte,
    posicao,
    equipe,
    treinador,
    altura,
    peso,
    imc,
    telefone,
    email,
    observacoes

):

    criar_atleta(nome)

    wb = abrir_atleta(nome)

    ws = wb["Cadastro"]

    ws["B4"] = nome
    ws["B5"] = sexo
    ws["B6"] = idade
    ws["B7"] = nascimento
    ws["B8"] = esporte
    ws["B9"] = posicao
    ws["B10"] = equipe
    ws["B11"] = treinador

    ws["B17"] = altura
    ws["B18"] = peso
    ws["B19"] = imc

    ws["B23"] = telefone
    ws["B24"] = email
    ws["B28"] = observacoes

    wb.save(caminho_atleta(nome))

    wb.close()

# ==========================================================
# LER CADASTRO
# ==========================================================

def ler_cadastro(nome):

    arquivo = caminho_atleta(nome)

    if not arquivo.exists():

        return None

    wb = load_workbook(

        arquivo,

        data_only=True

    )

    ws = wb["Cadastro"]

    dados = {

        "nome": ws["B4"].value or "",

        "sexo": ws["B5"].value or "",

        "idade": ws["B6"].value or 0,

        "nascimento": ws["B7"].value or "",

        "esporte": ws["B8"].value or "",

        "posicao": ws["B9"].value or "",

        "equipe": ws["B10"].value or "",

        "treinador": ws["B11"].value or "",

        "altura": ws["B17"].value or 0,

        "peso": ws["B18"].value or 0,

        "imc": ws["B19"].value or 0,

        "telefone": ws["B23"].value or "",

        "email": ws["B24"].value or "",

        "observacoes": ws["B28"].value or ""

    }

    wb.close()

    return dados
# ==========================================================
# ATUALIZAR CADASTRO
# ==========================================================

def atualizar_cadastro(nome, dados):

    wb = abrir_atleta(nome)

    if wb is None:

        return False

    ws = wb["Cadastro"]

    mapa = {

        "B4": "nome",
        "B5": "sexo",
        "B6": "idade",
        "B7": "nascimento",
        "B8": "esporte",
        "B9": "posicao",
        "B10": "equipe",
        "B11": "treinador",

        "B17": "altura",
        "B18": "peso",
        "B19": "imc",

        "B23": "telefone",
        "B24": "email",
        "B28": "observacoes"

    }

    for celula, chave in mapa.items():

        if chave in dados:

            ws[celula] = dados[chave]

    wb.save(caminho_atleta(nome))

    wb.close()

    return True


# ==========================================================
# DADOS DO ATLETA
# ==========================================================

def abrir_dados_atleta(nome):

    return ler_cadastro(nome)


# ==========================================================
# EXCLUIR
# ==========================================================

def excluir_atleta(nome):

    arquivo = caminho_atleta(nome)

    if arquivo.exists():

        arquivo.unlink()

        return True

    return False


# ==========================================================
# LISTAR TODOS
# ==========================================================

def listar_dados_atletas():

    atletas = []

    for nome in listar_atletas():

        try:

            dados = ler_cadastro(nome)

            if dados:

                atletas.append(dados)

        except Exception:

            pass

    return atletas


# ==========================================================
# PESQUISA
# ==========================================================

def pesquisar_atletas(texto):

    texto = (texto or "").lower().strip()

    resultado = []

    for atleta in listar_dados_atletas():

        nome = str(atleta.get("nome", "")).lower()

        esporte = str(atleta.get("esporte", "")).lower()

        equipe = str(atleta.get("equipe", "")).lower()

        treinador = str(atleta.get("treinador", "")).lower()

        if (

            texto in nome

            or texto in esporte

            or texto in equipe

            or texto in treinador

        ):

            resultado.append(atleta)

    return resultado


# ==========================================================
# ESTATÍSTICAS
# ==========================================================

def total_atletas():

    return len(listar_dados_atletas())


def ultimo_atleta():

    atletas = listar_dados_atletas()

    if not atletas:

        return None

    return atletas[-1]["nome"]


def media_idade():

    atletas = listar_dados_atletas()

    if not atletas:

        return 0

    soma = 0

    for atleta in atletas:

        soma += atleta.get("idade") or 0

    return round(soma / len(atletas), 1)


def media_imc():

    atletas = listar_dados_atletas()

    if not atletas:

        return 0

    soma = 0

    for atleta in atletas:

        soma += atleta.get("imc") or 0

    return round(soma / len(atletas), 2)
# ==========================================================
# ESPORTES
# ==========================================================

def esportes():

    contador = {}

    for atleta in listar_dados_atletas():

        esporte = atleta.get("esporte") or "Não informado"

        contador[esporte] = contador.get(esporte, 0) + 1

    return contador


# ==========================================================
# BUSCAR ATLETA PELO NOME
# ==========================================================

def atleta_por_nome(nome):

    return ler_cadastro(nome)


# ==========================================================
# DADOS PARA RELATÓRIO
# ==========================================================

def dados_relatorio(nome):

    dados = ler_cadastro(nome)

    if dados is None:

        return {}

    return dados


# ==========================================================
# RESUMO DO SISTEMA
# ==========================================================

def resumo():

    atletas = listar_dados_atletas()

    return {

        "total": len(atletas),

        "ultimo": ultimo_atleta(),

        "idade_media": media_idade(),

        "imc_medio": media_imc(),

        "esportes": esportes(),

        "atletas": atletas

    }


# ==========================================================
# EXPORTAÇÃO
# ==========================================================

def exportar_dados():

    return listar_dados_atletas()


# ==========================================================
# VERIFICAÇÃO
# ==========================================================

def sistema_ok():

    return {

        "modelo": MODELO.exists(),

        "pasta_modelos": PASTA_MODELOS.exists(),

        "pasta_atletas": PASTA_ATLETAS.exists(),

        "total_atletas": total_atletas()

    }


# ==========================================================
# LIMPAR PASTA DE ATLETAS
# ==========================================================

def limpar_atletas():

    total = 0

    for arquivo in PASTA_ATLETAS.glob("*.xlsx"):

        try:

            arquivo.unlink()

            total += 1

        except Exception:

            pass

    return total


# ==========================================================
# INFORMAÇÕES DO SISTEMA
# ==========================================================

def info_sistema():

    return {

        "base_dir": str(BASE_DIR),

        "modelo": str(MODELO),

        "pasta_atletas": str(PASTA_ATLETAS),

        "total": total_atletas()

    }


# ==========================================================
# FIM DO ARQUIVO
# ==========================================================
