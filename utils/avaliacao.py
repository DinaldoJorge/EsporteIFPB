# ==========================================================
# SPORTAVALIA PRO V2
# utils/avaliacao.py
# PARTE 1
# ==========================================================

from openpyxl import load_workbook

from utils.excel import caminho_atleta


# ==========================================================
# ABRIR PLANILHA
# ==========================================================

def abrir_avaliacao(nome):

    arquivo = caminho_atleta(nome)

    if not arquivo.exists():

        return None

    wb = load_workbook(arquivo)

    if "Avaliacao" not in wb.sheetnames:

        ws = wb.create_sheet("Avaliacao")

        ws["A1"] = "AVALIAÇÃO FÍSICA"
        ws["A4"] = "Velocidade"
        ws["A5"] = "Força"
        ws["A6"] = "Resistência"
        ws["A7"] = "Flexibilidade"
        ws["A8"] = "Coordenação"
        ws["A9"] = "Equilíbrio"
        ws["A12"] = "Observações"

        wb.save(arquivo)

    return wb


# ==========================================================
# LER AVALIAÇÃO
# ==========================================================

def ler_avaliacao(nome):

    wb = abrir_avaliacao(nome)

    if wb is None:

        return None

    ws = wb["Avaliacao"]

    dados = {

        "velocidade": ws["B4"].value,

        "forca": ws["B5"].value,

        "resistencia": ws["B6"].value,

        "flexibilidade": ws["B7"].value,

        "coordenacao": ws["B8"].value,

        "equilibrio": ws["B9"].value,

        "observacoes": ws["B12"].value

    }

    wb.close()

    return dados


# ==========================================================
# EXISTE AVALIAÇÃO?
# ==========================================================

def possui_avaliacao(nome):

    dados = ler_avaliacao(nome)

    if dados is None:

        return False

    return dados["velocidade"] is not None
# ==========================================================
# SALVAR AVALIAÇÃO
# ==========================================================

def salvar_avaliacao(
    nome,
    velocidade,
    forca,
    resistencia,
    flexibilidade,
    coordenacao,
    equilibrio,
    observacoes
):

    wb = abrir_avaliacao(nome)

    if wb is None:

        return False

    ws = wb["Avaliacao"]

    ws["B4"] = velocidade
    ws["B5"] = forca
    ws["B6"] = resistencia
    ws["B7"] = flexibilidade
    ws["B8"] = coordenacao
    ws["B9"] = equilibrio

    ws["B12"] = observacoes

    wb.save(caminho_atleta(nome))

    wb.close()

    return True


# ==========================================================
# MÉDIA DA AVALIAÇÃO
# ==========================================================

def calcular_media(nome):

    dados = ler_avaliacao(nome)

    if dados is None:

        return 0

    valores = [

        dados["velocidade"],
        dados["forca"],
        dados["resistencia"],
        dados["flexibilidade"],
        dados["coordenacao"],
        dados["equilibrio"]

    ]

    numeros = []

    for valor in valores:

        if valor is not None:

            numeros.append(float(valor))

    if len(numeros) == 0:

        return 0

    return round(sum(numeros) / len(numeros), 2)


# ==========================================================
# TOTAL DE PONTOS
# ==========================================================

def total_pontos(nome):

    dados = ler_avaliacao(nome)

    if dados is None:

        return 0

    total = 0

    for chave in [

        "velocidade",
        "forca",
        "resistencia",
        "flexibilidade",
        "coordenacao",
        "equilibrio"

    ]:

        valor = dados[chave]

        if valor is not None:

            total += float(valor)

    return round(total, 2)
# ==========================================================
# CLASSIFICAÇÃO
# ==========================================================

def classificar_desempenho(media):

    if media >= 9:

        return "🏆 Excelente"

    elif media >= 8:

        return "🥇 Muito Bom"

    elif media >= 7:

        return "🥈 Bom"

    elif media >= 6:

        return "🥉 Regular"

    elif media >= 5:

        return "⚠ Em desenvolvimento"

    else:

        return "❌ Necessita treinamento"


# ==========================================================
# RESUMO
# ==========================================================

def resumo_avaliacao(nome):

    dados = ler_avaliacao(nome)

    if dados is None:

        return None

    media = calcular_media(nome)

    total = total_pontos(nome)

    return {

        "dados": dados,

        "media": media,

        "total": total,

        "classificacao": classificar_desempenho(media)

    }


# ==========================================================
# ÚLTIMA AVALIAÇÃO
# ==========================================================

def ultima_avaliacao(nome):

    resumo = resumo_avaliacao(nome)

    if resumo is None:

        return None

    return resumo["dados"]


# ==========================================================
# ESTATÍSTICAS
# ==========================================================

def estatisticas_avaliacao(nome):

    resumo = resumo_avaliacao(nome)

    if resumo is None:

        return None

    return {

        "Velocidade": resumo["dados"]["velocidade"],

        "Força": resumo["dados"]["forca"],

        "Resistência": resumo["dados"]["resistencia"],

        "Flexibilidade": resumo["dados"]["flexibilidade"],

        "Coordenação": resumo["dados"]["coordenacao"],

        "Equilíbrio": resumo["dados"]["equilibrio"]

    }


# ==========================================================
# DASHBOARD DA AVALIAÇÃO
# ==========================================================

def indicadores_avaliacao(nome):

    resumo = resumo_avaliacao(nome)

    if resumo is None:

        return {

            "media": 0,

            "total": 0,

            "classificacao": "Sem avaliação"

        }

    return {

        "media": resumo["media"],

        "total": resumo["total"],

        "classificacao": resumo["classificacao"]

    }


# ==========================================================
# FIM DO ARQUIVO
# ==========================================================